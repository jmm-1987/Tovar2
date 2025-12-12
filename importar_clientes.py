"""Script para importar clientes desde Excel (ejecutar una sola vez)"""
import os
from datetime import datetime
from openpyxl import load_workbook

# Importar Flask app para tener acceso a la base de datos
from app import app, db
from models import Cliente

def importar_clientes():
    """Importar clientes desde el archivo Excel"""
    excel_path = os.path.join('static', 'clientes.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"Error: No se encontró el archivo {excel_path}")
        return
    
    print(f"Leyendo archivo: {excel_path}")
    
    with app.app_context():
        try:
            # Cargar el archivo Excel
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            # Leer la primera fila para obtener los encabezados
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value else '')
            
            print(f"Total de columnas: {len(headers)}")
            print(f"Encabezados: {headers}")
            
            # Mapeo de columnas del Excel a campos de la base de datos
            column_map = {}
            for idx, header in enumerate(headers):
                if header and header.strip():
                    header_upper = header.upper()
                    if 'NOMBRE FISCAL' in header_upper or ('NOMBRE' in header_upper and 'FISCAL' in header_upper):
                        column_map['nombre'] = idx
                        print(f"  [OK] Columna '{header}' -> nombre (indice {idx})")
                    elif 'ALIAS' in header_upper:
                        column_map['alias'] = idx
                        print(f"  [OK] Columna '{header}' -> alias (indice {idx})")
                    elif ('TEL' in header_upper or 'TELEFONO' in header_upper) and 'MOVIL' not in header_upper:
                        column_map['telefono'] = idx
                        print(f"  [OK] Columna '{header}' -> telefono (indice {idx})")
                    elif 'MOVIL' in header_upper or ('M' in header_upper and 'VIL' in header_upper):
                        column_map['movil'] = idx
                        print(f"  [OK] Columna '{header}' -> movil (indice {idx})")
                    elif 'E-MAIL' in header_upper or 'EMAIL' in header_upper:
                        column_map['email'] = idx
                        print(f"  [OK] Columna '{header}' -> email (indice {idx})")
                    elif 'PERSONA' in header_upper and 'CONTACTO' in header_upper:
                        column_map['personas_contacto'] = idx
                        print(f"  [OK] Columna '{header}' -> personas_contacto (indice {idx})")
                    elif 'N.I.F' in header_upper or ('NIF' in header_upper and '.' in header):
                        column_map['nif'] = idx
                        print(f"  [OK] Columna '{header}' -> nif (indice {idx})")
                    elif 'DOMICILIO' in header_upper or 'DIRECCION' in header_upper:
                        column_map['direccion'] = idx
                        print(f"  [OK] Columna '{header}' -> direccion (indice {idx})")
                    elif 'POBLACI' in header_upper:
                        column_map['poblacion'] = idx
                        print(f"  [OK] Columna '{header}' -> poblacion (indice {idx})")
                    elif ('C' in header_upper or 'COD' in header_upper) and 'POSTAL' in header_upper:
                        column_map['codigo_postal'] = idx
                        print(f"  [OK] Columna '{header}' -> codigo_postal (indice {idx})")
                    elif 'PROVINCIA' in header_upper:
                        column_map['provincia'] = idx
                        print(f"  [OK] Columna '{header}' -> provincia (indice {idx})")
                    elif 'ANOTACIONES' in header_upper:
                        column_map['anotaciones'] = idx
                        print(f"  [OK] Columna '{header}' -> anotaciones (indice {idx})")
            
            print(f"\n[OK] Mapa de columnas encontrado: {list(column_map.keys())}")
            
            # Verificar que tenemos al menos nombre
            if 'nombre' not in column_map:
                print("[ERROR] No se encontro la columna 'NOMBRE FISCAL'")
                return
            
            # Leer datos desde la fila 2 en adelante
            clientes_importados = 0
            clientes_duplicados = 0
            errores = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Obtener valores de las celdas según el mapeo
                    valores = {}
                    for campo, col_idx in column_map.items():
                        if col_idx < len(row):
                            valor = row[col_idx].value
                            if valor is not None and str(valor).strip() != '':
                                # Convertir a string y limpiar
                                if isinstance(valor, (int, float)):
                                    # Si es numérico, convertir a string
                                    if campo in ['telefono', 'movil', 'codigo_postal']:
                                        valor = str(int(valor))
                                    else:
                                        valor = str(int(valor))
                                else:
                                    valor = str(valor).strip()
                                
                                # Convertir a mayúsculas excepto email
                                if campo != 'email' and valor:
                                    valor = valor.upper()
                                
                                # Procesar código postal: añadir 0 si tiene 4 dígitos
                                if campo == 'codigo_postal' and valor:
                                    valor_str = str(valor).strip()
                                    # Eliminar espacios y caracteres no numéricos
                                    valor_str = ''.join(filter(str.isdigit, valor_str))
                                    if valor_str.isdigit() and len(valor_str) == 4:
                                        valor = '0' + valor_str
                                
                                valores[campo] = valor if valor else None
                            else:
                                valores[campo] = None
                        else:
                            valores[campo] = None
                    
                    # Verificar que al menos tenga nombre
                    if not valores.get('nombre'):
                        continue
                    
                    # Verificar si el cliente ya existe (por NIF o nombre)
                    cliente_existente = None
                    if valores.get('nif'):
                        cliente_existente = Cliente.query.filter_by(nif=valores['nif']).first()
                    
                    if not cliente_existente:
                        cliente_existente = Cliente.query.filter_by(nombre=valores['nombre']).first()
                    
                    if cliente_existente:
                        clientes_duplicados += 1
                        if clientes_duplicados <= 5:  # Mostrar solo los primeros 5
                            print(f"  Fila {row_idx}: Cliente '{valores['nombre']}' ya existe, se omite")
                        continue
                    
                    # Preparar email en minúsculas
                    email = valores.get('email')
                    if email:
                        email = email.lower()
                    
                    # Crear nuevo cliente
                    cliente = Cliente(
                        nombre=valores.get('nombre'),
                        alias=valores.get('alias'),
                        nif=valores.get('nif'),
                        direccion=valores.get('direccion'),
                        poblacion=valores.get('poblacion'),
                        provincia=valores.get('provincia'),
                        codigo_postal=valores.get('codigo_postal'),
                        pais='España',
                        telefono=valores.get('telefono'),
                        movil=valores.get('movil'),
                        email=email,
                        personas_contacto=valores.get('personas_contacto'),
                        anotaciones=valores.get('anotaciones'),
                        fecha_alta=datetime.now().date()
                    )
                    
                    db.session.add(cliente)
                    clientes_importados += 1
                    
                    if clientes_importados % 50 == 0:
                        db.session.commit()
                        print(f"  Procesados {clientes_importados} clientes...")
                        
                except Exception as e:
                    print(f"  [ERROR] Error en fila {row_idx}: {str(e)}")
                    errores += 1
                    continue
            
            # Hacer commit final de todos los clientes
            db.session.commit()
            print(f"\n{'='*60}")
            print(f"[OK] IMPORTACION COMPLETADA")
            print(f"{'='*60}")
            print(f"   - Clientes importados: {clientes_importados}")
            print(f"   - Clientes duplicados (omitidos): {clientes_duplicados}")
            print(f"   - Errores: {errores}")
            print(f"{'='*60}")
            
        except Exception as e:
            db.session.rollback()
            print(f"\n[ERROR] Error general: {str(e)}")
            import traceback
            traceback.print_exc()

if __name__ == '__main__':
    print("=" * 60)
    print("IMPORTACIÓN DE CLIENTES DESDE EXCEL")
    print("=" * 60)
    importar_clientes()
