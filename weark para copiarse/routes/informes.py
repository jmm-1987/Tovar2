"""Rutas para informes y reportes"""
from flask import Blueprint, render_template, request, jsonify, send_file, make_response, redirect, url_for
from flask_login import login_required
from datetime import datetime
from decimal import Decimal
from extensions import db
from models import Factura, FacturaProveedor, Nomina, Empleado, LineaFactura, Cliente, Ticket, OtroGasto, Presupuesto
from sqlalchemy import func, extract, not_
from sqlalchemy.orm import joinedload
from utils.auth import not_usuario_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from playwright.sync_api import sync_playwright
import io
import tempfile
import os

informes_bp = Blueprint('informes', __name__)


def _calcular_totales_presupuesto(presupuesto):
    base_imponible = Decimal('0.00')
    for linea in presupuesto.lineas or []:
        cantidad = Decimal(str(linea.cantidad)) if linea.cantidad else Decimal('0')
        precio_unit = Decimal(str(linea.precio_unitario)) if linea.precio_unitario else Decimal('0.00')
        descuento = Decimal(str(linea.descuento)) if linea.descuento else Decimal('0')

        precio_final = precio_unit
        if descuento > 0:
            if linea.precio_final:
                precio_final = Decimal(str(linea.precio_final))
            else:
                precio_final = precio_unit * (Decimal('1') - descuento / Decimal('100'))
        elif linea.precio_final:
            precio_final = Decimal(str(linea.precio_final))

        base_imponible += cantidad * precio_final

    base_q = base_imponible.quantize(Decimal('0.01'))
    if (presupuesto.estado or '') == 'cancelado':
        return base_q, Decimal('0.00'), base_q, Decimal('0')

    tipo_iva = Decimal('21')
    if presupuesto.cliente and getattr(presupuesto.cliente, 'tipo_iva', None) is not None:
        try:
            tipo_iva = Decimal(str(presupuesto.cliente.tipo_iva))
        except (ValueError, TypeError):
            pass

    iva_total = (base_imponible * tipo_iva / Decimal('100')).quantize(Decimal('0.01'))
    total = (base_imponible + iva_total).quantize(Decimal('0.01'))
    return base_q, iva_total, total, tipo_iva


def _obtener_presupuestos_emitidos_fd(año, tipo_filtro='mes', periodo=None, cliente_id=None):
    query = Presupuesto.query.options(
        joinedload(Presupuesto.cliente),
        joinedload(Presupuesto.lineas)
    ).filter(
        Presupuesto.fd_deliverynote_uuid.isnot(None)
    )
    if cliente_id:
        query = query.filter(Presupuesto.cliente_id == cliente_id)

    presupuestos = []
    for p in query.all():
        fecha_ref = p.fd_deliverynote_sent_at or p.fecha_creacion
        if not fecha_ref:
            continue
        if fecha_ref.year != año:
            continue
        if tipo_filtro == 'mes' and periodo and fecha_ref.month != periodo:
            continue
        if tipo_filtro == 'trimestre' and periodo:
            mes_inicio = (periodo - 1) * 3 + 1
            mes_fin = periodo * 3
            if not (mes_inicio <= fecha_ref.month <= mes_fin):
                continue
        presupuestos.append(p)

    presupuestos.sort(key=lambda x: ((x.fd_deliverynote_sent_at or x.fecha_creacion), (x.numero_solicitud or '')), reverse=True)
    return presupuestos

@informes_bp.route('/informes')
@login_required
@not_usuario_required
def index():
    """Página principal de informes"""
    return render_template('informes/index.html')


MESES_NOMBRES = [
    '',
    'Enero',
    'Febrero',
    'Marzo',
    'Abril',
    'Mayo',
    'Junio',
    'Julio',
    'Agosto',
    'Septiembre',
    'Octubre',
    'Noviembre',
    'Diciembre',
]


@informes_bp.route('/informes/presupuestos')
@login_required
@not_usuario_required
def informes_presupuestos():
    """Informe de presupuestos: solo carga datos tras pulsar Mostrar (mostrar=1)."""
    mostrar_resultados = request.args.get('mostrar', type=int) == 1
    año = request.args.get('año', datetime.now().year, type=int)
    mes = request.args.get('mes', None, type=int)
    if mes is not None and (mes < 1 or mes > 12):
        mes = None
    cliente_id = request.args.get('cliente_id', None, type=int)
    fd_filtro = (request.args.get('fd') or '').strip().lower()
    if fd_filtro not in ('', 'si', 'no'):
        fd_filtro = ''

    filas = []
    fd_enviados = {'n': 0, 'base': Decimal('0'), 'iva': Decimal('0'), 'total': Decimal('0')}
    fd_no_enviados = {'n': 0, 'base': Decimal('0'), 'iva': Decimal('0'), 'total': Decimal('0')}

    if mostrar_resultados:
        q = Presupuesto.query.options(
            joinedload(Presupuesto.cliente),
            joinedload(Presupuesto.lineas),
        ).filter(extract('year', Presupuesto.fecha_creacion) == año)
        if mes:
            q = q.filter(extract('month', Presupuesto.fecha_creacion) == mes)
        if cliente_id:
            q = q.filter(Presupuesto.cliente_id == cliente_id)

        presupuestos = q.order_by(Presupuesto.fecha_creacion.desc()).all()

        for p in presupuestos:
            base, iva, tot, tipo_iva = _calcular_totales_presupuesto(p)
            en_fd = bool(p.fd_deliverynote_uuid)
            if en_fd:
                fd_enviados['n'] += 1
                fd_enviados['base'] += base
                fd_enviados['iva'] += iva
                fd_enviados['total'] += tot
            else:
                fd_no_enviados['n'] += 1
                fd_no_enviados['base'] += base
                fd_no_enviados['iva'] += iva
                fd_no_enviados['total'] += tot

            if fd_filtro == 'si' and not en_fd:
                continue
            if fd_filtro == 'no' and en_fd:
                continue

            fc = p.fecha_creacion
            filas.append(
                {
                    'id': p.id,
                    'numero': p.numero_solicitud or str(p.id),
                    'fecha_creacion': fc,
                    'cliente': p.cliente.nombre if p.cliente else 'N/A',
                    'base': base,
                    'iva': iva,
                    'total': tot,
                    'tipo_iva': tipo_iva,
                    'fd': en_fd,
                }
            )

    if mes:
        periodo_label = f'{MESES_NOMBRES[mes]} {año}'
    else:
        periodo_label = f'Año {año}'

    clientes = Cliente.query.order_by(Cliente.nombre).all()
    cliente_seleccionado = Cliente.query.get(cliente_id) if cliente_id else None

    y_now = datetime.now().year
    años_opciones = list(range(y_now - 10, y_now + 5))
    if año not in años_opciones:
        años_opciones.append(año)
        años_opciones.sort()

    return render_template(
        'informes/presupuestos.html',
        filas=filas,
        meses_nombres=MESES_NOMBRES,
        años_opciones=años_opciones,
        año=año,
        mes=mes,
        periodo_label=periodo_label,
        clientes=clientes,
        cliente_id=cliente_id,
        cliente_seleccionado=cliente_seleccionado,
        fd_filtro=fd_filtro,
        fd_enviados=fd_enviados,
        fd_no_enviados=fd_no_enviados,
        mostrar_resultados=mostrar_resultados,
    )


@informes_bp.route('/informes/facturacion-emitida')
@login_required
@not_usuario_required
def facturacion_emitida():
    """Informe de emisión basado en presupuestos enviados a FD."""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')  # 'mes' o 'trimestre'
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    presupuestos = _obtener_presupuestos_emitidos_fd(año, tipo_filtro, periodo, cliente_id)
    if tipo_filtro == 'mes' and periodo:
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"

    filas = []
    total_facturacion = Decimal('0.00')
    total_iva_repercutido = Decimal('0.00')
    for p in presupuestos:
        base, iva, total, tipo_iva = _calcular_totales_presupuesto(p)
        total_facturacion += total
        total_iva_repercutido += iva
        fecha_ref = p.fd_deliverynote_sent_at or p.fecha_creacion
        filas.append({
            'id': p.id,
            'numero': p.numero_solicitud or str(p.id),
            'fecha': fecha_ref,
            'nombre': p.cliente.nombre if p.cliente else 'N/A',
            'nif': p.cliente.nif if p.cliente else '',
            'base': base,
            'iva': iva,
            'total': total,
            'tipo_iva': tipo_iva,
            'fd_deliverynote_uuid': p.fd_deliverynote_uuid,
            'fd_deliverynote_doc_number': p.fd_deliverynote_doc_number
        })
    
    # Obtener lista de clientes para el filtro
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    cliente_seleccionado = Cliente.query.get(cliente_id) if cliente_id else None
    
    return render_template('informes/facturacion_emitida.html', 
                         facturas=filas,
                         total_facturacion=total_facturacion,
                         total_iva_repercutido=total_iva_repercutido,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         clientes=clientes,
                         cliente_id=cliente_id,
                         cliente_seleccionado=cliente_seleccionado)

@informes_bp.route('/informes/facturacion-emitida/exportar-excel')
@login_required
@not_usuario_required
def facturacion_emitida_exportar_excel():
    """Exportar informe de presupuestos enviados a FD a Excel."""
    # Obtener parámetros de filtro (mismos que en el informe principal)
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    presupuestos = _obtener_presupuestos_emitidos_fd(año, tipo_filtro, periodo, cliente_id)
    if tipo_filtro == 'mes' and periodo:
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos Enviados FD"
    
    # Estilos para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    headers = ['DNI/CIF', 'Nombre Cliente', 'Número Presupuesto', 'Fecha', 'Base Imponible', 'IVA', 'Total', 'Estado']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Escribir datos
    for row_idx, p in enumerate(presupuestos, 2):
        base_imponible, importe_iva, total, _tipo = _calcular_totales_presupuesto(p)
        fecha_ref = p.fd_deliverynote_sent_at or p.fecha_creacion
        dni_cif = p.cliente.nif if p.cliente and p.cliente.nif else ''
        nombre_cliente = p.cliente.nombre if p.cliente and p.cliente.nombre else ''
        numero_presupuesto = p.numero_solicitud or str(p.id)
        fecha_expedicion = fecha_ref.strftime('%d/%m/%Y') if fecha_ref else ''
        
        # Escribir fila
        ws.cell(row=row_idx, column=1, value=dni_cif)
        ws.cell(row=row_idx, column=2, value=nombre_cliente)
        ws.cell(row=row_idx, column=3, value=numero_presupuesto)
        ws.cell(row=row_idx, column=4, value=fecha_expedicion)
        ws.cell(row=row_idx, column=5, value=round(float(base_imponible), 2))
        ws.cell(row=row_idx, column=6, value=round(float(importe_iva), 2))
        ws.cell(row=row_idx, column=7, value=round(float(total), 2))
        ws.cell(row=row_idx, column=8, value='Enviado FD')
        
        # Formato numérico para columnas de importes
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'
    
    # Ajustar ancho de columnas automáticamente según el contenido
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # Para números, calcular longitud considerando formato
                    if isinstance(cell.value, (int, float)):
                        cell_length = len(f"{cell.value:,.2f}")
                    else:
                        cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        # Añadir un poco de espacio extra y establecer límites mínimos y máximos
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Ajustar altura de fila de encabezado
    ws.row_dimensions[1].height = 25
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo con el período
    filename = f'facturacion_emitida_{periodo_label.replace(" ", "_")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@informes_bp.route('/informes/facturacion-emitida/exportar-pdf')
@login_required
@not_usuario_required
def facturacion_emitida_exportar_pdf():
    """Exportar informe de presupuestos enviados a FD a PDF."""
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    presupuestos = _obtener_presupuestos_emitidos_fd(año, tipo_filtro, periodo, cliente_id)
    
    if tipo_filtro == 'mes' and periodo:
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"

    filas = []
    total_facturacion = Decimal('0')
    total_base = Decimal('0')
    total_iva = Decimal('0')
    for p in presupuestos:
        base, iva, total, _tipo = _calcular_totales_presupuesto(p)
        total_base += base
        total_iva += iva
        total_facturacion += total
        fecha_ref = p.fd_deliverynote_sent_at or p.fecha_creacion
        filas.append({
            'numero': p.numero_solicitud or str(p.id),
            'fecha': fecha_ref,
            'nombre': p.cliente.nombre if p.cliente else '',
            'nif': p.cliente.nif if p.cliente else '',
            'base': base,
            'iva': iva,
            'total': total
        })
    
    total_iva_repercutido = total_iva
    
    html = render_template('informes/facturacion_emitida_pdf.html',
        facturas=filas,
        periodo_label=periodo_label,
        total_facturacion=total_facturacion,
        total_iva_repercutido=total_iva_repercutido,
        total_base=total_base,
        total_iva=total_iva
    )
    
    pdf_buffer = BytesIO()
    try:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as temp_file:
            temp_file.write(html)
            temp_html_path = temp_file.name
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(f'file://{temp_html_path}')
            pdf_bytes = page.pdf(
                format='A4',
                print_background=True,
                margin={'top': '6mm', 'right': '6mm', 'bottom': '6mm', 'left': '6mm'}
            )
            browser.close()
        
        pdf_buffer.write(pdf_bytes)
        try:
            os.unlink(temp_html_path)
        except Exception:
            pass
    except Exception as e:
        import traceback
        from flask import flash
        print(traceback.format_exc())
        flash(f'Error al generar PDF: {str(e)}', 'error')
        return redirect(url_for('informes.facturacion_emitida'))
    
    pdf_buffer.seek(0)
    filename = f'facturacion_emitida_{periodo_label.replace(" ", "_")}.pdf'
    
    response = make_response(pdf_buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename={filename}'
    return response

@informes_bp.route('/informes/facturacion-soportada')
@login_required
@not_usuario_required
def facturacion_soportada():
    """Informe de facturación soportada (facturas de proveedor) con filtros por mes o trimestre"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')  # 'mes' o 'trimestre'
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    # Base query
    query = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(FacturaProveedor.fecha_factura.desc()).all()
    
    # Calcular totales
    total_facturacion = sum(f.total for f in facturas)
    total_base_imponible = sum(f.base_imponible for f in facturas)
    total_iva_soportado = sum(f.importe_iva for f in facturas)
    
    return render_template('informes/facturacion_soportada.html',
                         facturas=facturas,
                         total_facturacion=total_facturacion,
                         total_base_imponible=total_base_imponible,
                         total_iva_soportado=total_iva_soportado,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label)

@informes_bp.route('/informes/nominas')
@login_required
@not_usuario_required
def nominas():
    """Informe de nóminas por empleado y global"""
    # Obtener parámetros de filtro
    empleado_id = request.args.get('empleado_id', None, type=int)
    año = request.args.get('año', datetime.now().year, type=int)
    
    # Base query
    query = Nomina.query.filter(Nomina.año == año)
    
    if empleado_id:
        query = query.filter(Nomina.empleado_id == empleado_id)
    
    nominas = query.order_by(Nomina.mes.desc(), Nomina.empleado_id).all()
    
    # Calcular totales
    total_global = sum(n.total_devengado for n in nominas)
    
    # Totales por empleado
    totales_por_empleado = {}
    for nomina in nominas:
        empleado_nombre = nomina.empleado.nombre if nomina.empleado else 'Sin empleado'
        if empleado_nombre not in totales_por_empleado:
            totales_por_empleado[empleado_nombre] = Decimal('0')
        totales_por_empleado[empleado_nombre] += nomina.total_devengado
    
    # Obtener lista de empleados para el filtro
    empleados = Empleado.query.order_by(Empleado.nombre).all()
    
    return render_template('informes/nominas.html',
                         nominas=nominas,
                         total_global=total_global,
                         totales_por_empleado=totales_por_empleado,
                         empleados=empleados,
                         empleado_id=empleado_id,
                         año=año)

@informes_bp.route('/informes/iva')
@login_required
@not_usuario_required
def iva():
    """Informe de IVA: contrastar IVA repercutido vs IVA soportado"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')  # 'mes' o 'trimestre'
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    presupuestos_emitidos = _obtener_presupuestos_emitidos_fd(año, tipo_filtro, periodo)
    if tipo_filtro == 'mes' and periodo:
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"

    iva_repercutido = Decimal('0')
    base_repercutida = Decimal('0')
    for p in presupuestos_emitidos:
        base, iva, _total, _tipo_iva = _calcular_totales_presupuesto(p)
        base_repercutida += base
        iva_repercutido += iva
    
    # Calcular IVA soportado (de facturas de proveedor)
    query_facturas_prov = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
    
    facturas_proveedor = query_facturas_prov.all()
    iva_soportado = sum(f.importe_iva for f in facturas_proveedor)
    base_soportada = sum(f.base_imponible for f in facturas_proveedor)
    
    # Calcular diferencia
    diferencia_iva = iva_repercutido - iva_soportado
    
    return render_template('informes/iva.html',
                         iva_repercutido=iva_repercutido,
                         base_repercutida=base_repercutida,
                         iva_soportado=iva_soportado,
                         base_soportada=base_soportada,
                         diferencia_iva=diferencia_iva,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         num_facturas_emitidas=len(presupuestos_emitidos),
                         num_facturas_proveedor=len(facturas_proveedor))

@informes_bp.route('/informes/resultados')
@login_required
@not_usuario_required
def resultados():
    """Informe de resultados: ingresos - gastos categorizados por tipo, con gráficas mes a mes"""
    año = request.args.get('año', datetime.now().year, type=int)
    
    MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 
             'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    datos_meses = []
    
    for mes in range(1, 13):
        # Ingresos: presupuestos enviados FD + tickets tienda
        presupuestos_mes = _obtener_presupuestos_emitidos_fd(año, 'mes', mes, None)
        ingresos_facturas = Decimal('0')
        for p in presupuestos_mes:
            _base, _iva, total, _tipo = _calcular_totales_presupuesto(p)
            ingresos_facturas += total
        
        tickets = Ticket.query.filter(
            extract('year', Ticket.fecha_expedicion) == año,
            extract('month', Ticket.fecha_expedicion) == mes
        ).all()
        ingresos_tickets = sum(Decimal(str(t.importe_total)) for t in tickets)
        ingresos = ingresos_facturas + ingresos_tickets
        
        # Gastos por categoría
        facturas_prov = FacturaProveedor.query.filter(
            extract('year', FacturaProveedor.fecha_factura) == año,
            extract('month', FacturaProveedor.fecha_factura) == mes
        ).all()
        gasto_facturas = sum(Decimal(str(f.total)) for f in facturas_prov)
        
        nominas = Nomina.query.filter(Nomina.año == año, Nomina.mes == mes).all()
        gasto_nominas = sum(Decimal(str(n.total_devengado)) for n in nominas)
        
        recibo_iva = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'recibo_iva'
        ).all()
        gasto_iva = sum(Decimal(str(g.importe)) for g in recibo_iva)
        
        recibo_irpf = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'recibo_irpf'
        ).all()
        gasto_irpf = sum(Decimal(str(g.importe)) for g in recibo_irpf)
        
        seg_social = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'seguridad_social'
        ).all()
        gasto_seg_social = sum(Decimal(str(g.importe)) for g in seg_social)
        
        total_gastos = gasto_facturas + gasto_nominas + gasto_iva + gasto_irpf + gasto_seg_social
        resultado = ingresos - total_gastos
        
        datos_meses.append({
            'mes': mes,
            'mes_nombre': MESES[mes],
            'ingresos': float(ingresos),
            'gasto_facturas': float(gasto_facturas),
            'gasto_nominas': float(gasto_nominas),
            'gasto_iva': float(gasto_iva),
            'gasto_irpf': float(gasto_irpf),
            'gasto_seg_social': float(gasto_seg_social),
            'total_gastos': float(total_gastos),
            'resultado': float(resultado)
        })
    
    # Totales anuales
    total_ingresos = sum(d['ingresos'] for d in datos_meses)
    total_gastos_anual = sum(d['total_gastos'] for d in datos_meses)
    resultado_anual = total_ingresos - total_gastos_anual
    
    return render_template('informes/resultados.html',
                         datos_meses=datos_meses,
                         año=año,
                         total_ingresos=total_ingresos,
                         total_gastos_anual=total_gastos_anual,
                         resultado_anual=resultado_anual)


@informes_bp.route('/informes/resultados-nuevo')
@login_required
@not_usuario_required
def resultados_nuevo():
    """Informe de resultados usando presupuestos (FD enviados y opcional no enviados)."""
    año = request.args.get('año', datetime.now().year, type=int)
    mes_filtro = request.args.get('mes', type=int)
    if mes_filtro is not None and (mes_filtro < 1 or mes_filtro > 12):
        mes_filtro = None

    MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
             'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    datos_meses = []
    meses_a_procesar = [mes_filtro] if mes_filtro else list(range(1, 13))

    for mes in meses_a_procesar:
        presupuestos_mes = Presupuesto.query.options(
            joinedload(Presupuesto.lineas),
            joinedload(Presupuesto.cliente),
        ).filter(
            extract('year', Presupuesto.fecha_creacion) == año,
            extract('month', Presupuesto.fecha_creacion) == mes
        ).all()

        ingresos_presupuestos_fd = Decimal('0')
        ingresos_presupuestos_cancelado = Decimal('0')

        for p in presupuestos_mes:
            _base, _iva, total, _tipo = _calcular_totales_presupuesto(p)
            if p.fd_deliverynote_uuid:
                ingresos_presupuestos_fd += total
            if p.estado == 'cancelado':
                ingresos_presupuestos_cancelado += total

        ingresos_presupuestos_no_fd = Decimal('0')
        ingresos_presupuestos = ingresos_presupuestos_cancelado
        ingresos = ingresos_presupuestos

        facturas_prov = FacturaProveedor.query.filter(
            extract('year', FacturaProveedor.fecha_factura) == año,
            extract('month', FacturaProveedor.fecha_factura) == mes
        ).all()
        gasto_facturas = sum(Decimal(str(f.total)) for f in facturas_prov)

        nominas = Nomina.query.filter(Nomina.año == año, Nomina.mes == mes).all()
        gasto_nominas = sum(Decimal(str(n.total_devengado)) for n in nominas)

        recibo_iva = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'recibo_iva'
        ).all()
        gasto_iva = sum(Decimal(str(g.importe)) for g in recibo_iva)

        recibo_irpf = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'recibo_irpf'
        ).all()
        gasto_irpf = sum(Decimal(str(g.importe)) for g in recibo_irpf)

        seg_social = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'seguridad_social'
        ).all()
        gasto_seg_social = sum(Decimal(str(g.importe)) for g in seg_social)

        total_gastos = gasto_facturas + gasto_nominas + gasto_iva + gasto_irpf + gasto_seg_social
        # El resultado económico se calcula con lo efectivamente enviado a FD.
        resultado = ingresos_presupuestos_fd - total_gastos

        datos_meses.append({
            'mes': mes,
            'mes_nombre': MESES[mes],
            'ingresos': float(ingresos),
            'ingresos_presupuestos': float(ingresos_presupuestos),
            'ingresos_presupuestos_fd': float(ingresos_presupuestos_fd),
            'ingresos_presupuestos_cancelado': float(ingresos_presupuestos_cancelado),
            'ingresos_presupuestos_no_fd': float(ingresos_presupuestos_no_fd),
            'gasto_facturas': float(gasto_facturas),
            'gasto_nominas': float(gasto_nominas),
            'gasto_iva': float(gasto_iva),
            'gasto_irpf': float(gasto_irpf),
            'gasto_seg_social': float(gasto_seg_social),
            'total_gastos': float(total_gastos),
            'resultado': float(resultado)
        })

    total_ingresos = sum(d['ingresos'] for d in datos_meses)
    total_ingresos_fd = sum(d['ingresos_presupuestos_fd'] for d in datos_meses)
    total_gastos_anual = sum(d['total_gastos'] for d in datos_meses)
    resultado_anual = total_ingresos_fd - total_gastos_anual
    total_presupuestos_fd = sum(d['ingresos_presupuestos_fd'] for d in datos_meses)
    total_presupuestos_cancelado = sum(d['ingresos_presupuestos_cancelado'] for d in datos_meses)
    total_presupuestos_no_fd = sum(d['ingresos_presupuestos_no_fd'] for d in datos_meses)

    return render_template(
        'informes/resultados_nuevo.html',
        datos_meses=datos_meses,
        año=año,
        mes_filtro=mes_filtro,
        mes_filtro_nombre=(MESES[mes_filtro] if mes_filtro else ''),
        total_ingresos=total_ingresos,
        total_ingresos_fd=total_ingresos_fd,
        total_gastos_anual=total_gastos_anual,
        resultado_anual=resultado_anual,
        total_presupuestos_fd=total_presupuestos_fd,
        total_presupuestos_cancelado=total_presupuestos_cancelado,
        total_presupuestos_no_fd=total_presupuestos_no_fd
    )


@informes_bp.route('/informes/resultados-nuevo/detalle')
@login_required
@not_usuario_required
def resultados_nuevo_detalle():
    """Detalle mensual por categoría del informe de resultados nuevo."""
    año = request.args.get('año', datetime.now().year, type=int)
    mes = request.args.get('mes', type=int)
    categoria = (request.args.get('categoria') or '').strip()
    if not mes or mes < 1 or mes > 12:
        return redirect(url_for('informes.resultados_nuevo', año=año))

    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
             'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    presupuestos_mes = Presupuesto.query.options(
        joinedload(Presupuesto.lineas),
        joinedload(Presupuesto.cliente)
    ).filter(
        extract('year', Presupuesto.fecha_creacion) == año,
        extract('month', Presupuesto.fecha_creacion) == mes
    ).all()

    presupuestos_fd = []
    presupuestos_cancelado = []
    presupuestos_no_fd = []
    for p in presupuestos_mes:
        _b, _i, total, _t = _calcular_totales_presupuesto(p)
        fila = {
            'fecha': p.fecha_creacion.strftime('%d/%m/%Y') if p.fecha_creacion else '',
            'referencia': p.numero_solicitud or str(p.id),
            'cliente': p.cliente.nombre if p.cliente else 'N/A',
            'detalle': 'Enviado a FD' if p.fd_deliverynote_uuid else 'No enviado a FD',
            'importe': float(total)
        }
        if p.fd_deliverynote_uuid:
            presupuestos_fd.append(fila)
        if p.estado == 'cancelado':
            presupuestos_cancelado.append({
                **fila,
                'detalle': (fila.get('detalle', '') + ' | Estado cancelado').strip(' |')
            })

    nominas = Nomina.query.options(joinedload(Nomina.empleado)).filter(
        Nomina.año == año,
        Nomina.mes == mes
    ).all()
    filas_nominas = [
        {
            'fecha': f'01/{mes:02d}/{año}',
            'referencia': str(n.id),
            'cliente': n.empleado.nombre if n.empleado else 'Sin empleado',
            'detalle': n.observaciones or '',
            'importe': float(n.total_devengado or 0)
        }
        for n in nominas
    ]

    facturas_prov = FacturaProveedor.query.options(joinedload(FacturaProveedor.proveedor)).filter(
        extract('year', FacturaProveedor.fecha_factura) == año,
        extract('month', FacturaProveedor.fecha_factura) == mes
    ).all()
    filas_facturas_prov = [
        {
            'fecha': f.fecha_factura.strftime('%d/%m/%Y') if f.fecha_factura else '',
            'referencia': f.numero_factura or str(f.id),
            'cliente': f.proveedor.nombre if f.proveedor else 'Sin proveedor',
            'detalle': f.observaciones or '',
            'importe': float(f.total or 0)
        }
        for f in facturas_prov
    ]

    def obtener_otros_gastos(tipo):
        gastos = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == tipo
        ).all()
        return [
            {
                'fecha': g.fecha.strftime('%d/%m/%Y') if g.fecha else '',
                'referencia': str(g.id),
                'cliente': '-',
                'detalle': g.observaciones or (g.periodo or ''),
                'importe': float(g.importe or 0)
            }
            for g in gastos
        ]

    filas_iva = obtener_otros_gastos('recibo_iva')
    filas_irpf = obtener_otros_gastos('recibo_irpf')
    filas_seg_social = obtener_otros_gastos('seguridad_social')

    tot_fd = sum(x['importe'] for x in presupuestos_fd)
    tot_cancelado = sum(x['importe'] for x in presupuestos_cancelado)
    tot_no_fd = sum(x['importe'] for x in presupuestos_no_fd)
    ingresos = tot_cancelado
    tot_facturas_prov = sum(x['importe'] for x in filas_facturas_prov)
    tot_nominas = sum(x['importe'] for x in filas_nominas)
    tot_iva = sum(x['importe'] for x in filas_iva)
    tot_irpf = sum(x['importe'] for x in filas_irpf)
    tot_seg_social = sum(x['importe'] for x in filas_seg_social)
    total_gastos = tot_facturas_prov + tot_nominas + tot_iva + tot_irpf + tot_seg_social
    resultado = tot_fd - total_gastos

    titulos = {
        'ingresos_totales': 'Detalle de ingresos considerados (estado cancelado)',
        'ingresos_fd': 'Detalle de presupuestos enviados FD',
        'ingresos_no_fd': 'Detalle de presupuestos no enviados FD',
        'gastos_totales': 'Detalle de gastos totales',
        'resultado': 'Detalle de resultado',
        'gasto_nominas': 'Detalle de nóminas',
        'gasto_facturas': 'Detalle de facturas de proveedor',
        'gasto_iva': 'Detalle de recibo IVA',
        'gasto_irpf': 'Detalle de recibo IRPF',
        'gasto_seg_social': 'Detalle de seguridad social'
    }
    titulo = titulos.get(categoria, 'Detalle mensual')

    secciones = []
    if categoria == 'ingresos_fd':
        secciones.append({'titulo': 'Presupuestos enviados FD', 'filas': presupuestos_fd, 'total': tot_fd})
    elif categoria == 'ingresos_no_fd':
        secciones.append({'titulo': 'Presupuestos no enviados FD', 'filas': [], 'total': 0})
    elif categoria == 'gasto_facturas':
        secciones.append({'titulo': 'Facturas de proveedor', 'filas': filas_facturas_prov, 'total': tot_facturas_prov})
    elif categoria == 'gasto_nominas':
        secciones.append({'titulo': 'Nóminas', 'filas': filas_nominas, 'total': tot_nominas})
    elif categoria == 'gasto_iva':
        secciones.append({'titulo': 'Recibo IVA', 'filas': filas_iva, 'total': tot_iva})
    elif categoria == 'gasto_irpf':
        secciones.append({'titulo': 'Recibo IRPF', 'filas': filas_irpf, 'total': tot_irpf})
    elif categoria == 'gasto_seg_social':
        secciones.append({'titulo': 'Seguridad social', 'filas': filas_seg_social, 'total': tot_seg_social})
    elif categoria == 'gastos_totales':
        secciones.extend([
            {'titulo': 'Facturas de proveedor', 'filas': filas_facturas_prov, 'total': tot_facturas_prov},
            {'titulo': 'Nóminas', 'filas': filas_nominas, 'total': tot_nominas},
            {'titulo': 'Recibo IVA', 'filas': filas_iva, 'total': tot_iva},
            {'titulo': 'Recibo IRPF', 'filas': filas_irpf, 'total': tot_irpf},
            {'titulo': 'Seguridad social', 'filas': filas_seg_social, 'total': tot_seg_social}
        ])
    elif categoria == 'resultado':
        secciones.extend([
            {'titulo': 'Presupuestos enviados FD', 'filas': presupuestos_fd, 'total': tot_fd},
            {'titulo': 'Presupuestos no enviados FD', 'filas': presupuestos_no_fd, 'total': tot_no_fd},
            {'titulo': 'Facturas de proveedor', 'filas': filas_facturas_prov, 'total': tot_facturas_prov},
            {'titulo': 'Nóminas', 'filas': filas_nominas, 'total': tot_nominas},
            {'titulo': 'Recibo IVA', 'filas': filas_iva, 'total': tot_iva},
            {'titulo': 'Recibo IRPF', 'filas': filas_irpf, 'total': tot_irpf},
            {'titulo': 'Seguridad social', 'filas': filas_seg_social, 'total': tot_seg_social}
        ])
    else:
        secciones.extend([
            {'titulo': 'Presupuestos considerados (estado cancelado)', 'filas': presupuestos_cancelado, 'total': tot_cancelado},
            {'titulo': 'Presupuestos enviados FD', 'filas': presupuestos_fd, 'total': tot_fd}
        ])

    return render_template(
        'informes/resultados_nuevo_detalle.html',
        año=año,
        mes=mes,
        mes_nombre=meses[mes],
        categoria=categoria,
        titulo=titulo,
        incluir_no_fd=False,
        secciones=secciones,
        resumen={
            'ingresos': ingresos,
            'ingresos_fd': tot_fd,
            'ingresos_no_fd': tot_no_fd,
            'total_gastos': total_gastos,
            'resultado': resultado
        }
    )

@informes_bp.route('/informes/facturacion-emitida/detalle')
@login_required
@not_usuario_required
def facturacion_emitida_detalle():
    """Detalle de emisión basado en presupuestos enviados a FD."""
    # Obtener parámetros de filtro (mismos que en el informe principal)
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    presupuestos = _obtener_presupuestos_emitidos_fd(año, tipo_filtro, periodo, cliente_id)
    if tipo_filtro == 'mes' and periodo:
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"

    facturas = []
    for p in presupuestos:
        base, iva, total, tipo_iva = _calcular_totales_presupuesto(p)
        fecha_ref = p.fd_deliverynote_sent_at or p.fecha_creacion
        facturas.append({
            'id': p.id,
            'serie': 'FD',
            'numero': p.numero_solicitud or str(p.id),
            'fecha_expedicion': fecha_ref,
            'nombre': p.cliente.nombre if p.cliente else 'N/A',
            'nif': p.cliente.nif if p.cliente else '',
            'importe_total': total,
            'tipo_iva': tipo_iva,
            'descuento_pronto_pago': 0,
            'estado_cobro': 'enviado_fd',
            'fd_deliverynote_uuid': p.fd_deliverynote_uuid,
            'fd_deliverynote_doc_number': p.fd_deliverynote_doc_number,
            'base_calc': base,
            'iva_calc': iva
        })
    
    # Obtener cliente seleccionado si existe
    cliente_seleccionado = Cliente.query.get(cliente_id) if cliente_id else None
    
    return render_template('informes/detalle_facturacion_emitida.html',
                         facturas=facturas,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         cliente_id=cliente_id,
                         cliente_seleccionado=cliente_seleccionado)

@informes_bp.route('/informes/facturacion-soportada/detalle')
@login_required
@not_usuario_required
def facturacion_soportada_detalle():
    """Detalle de facturación soportada con listado completo de facturas de proveedor"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    # Base query
    query = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(FacturaProveedor.fecha_factura.desc()).all()
    
    return render_template('informes/detalle_facturacion_soportada.html',
                         facturas=facturas,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label)

@informes_bp.route('/informes/nominas/detalle')
@login_required
@not_usuario_required
def nominas_detalle():
    """Detalle de nóminas con listado completo"""
    # Obtener parámetros de filtro
    empleado_id = request.args.get('empleado_id', None, type=int)
    año = request.args.get('año', datetime.now().year, type=int)
    
    # Base query
    query = Nomina.query.filter(Nomina.año == año)
    
    if empleado_id:
        query = query.filter(Nomina.empleado_id == empleado_id)
    
    nominas = query.order_by(Nomina.mes.desc(), Nomina.empleado_id).all()
    
    # Obtener lista de empleados para el filtro
    empleados = Empleado.query.order_by(Empleado.nombre).all()
    empleado_seleccionado = Empleado.query.get(empleado_id) if empleado_id else None
    
    return render_template('informes/detalle_nominas.html',
                         nominas=nominas,
                         empleados=empleados,
                         empleado_id=empleado_id,
                         empleado_seleccionado=empleado_seleccionado,
                         año=año)

@informes_bp.route('/informes/iva/detalle')
@login_required
@not_usuario_required
def iva_detalle():
    """Detalle de IVA con listado completo de facturas emitidas y de proveedor"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    presupuestos_emitidos = _obtener_presupuestos_emitidos_fd(año, tipo_filtro, periodo)
    if tipo_filtro == 'mes' and periodo:
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    # Calcular IVA soportado (de facturas de proveedor)
    query_facturas_prov = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
    
    facturas_proveedor = query_facturas_prov.order_by(FacturaProveedor.fecha_factura.desc()).all()
    
    # Precalcular totales para evitar problemas de tipo en las plantillas
    total_base_emitida = Decimal('0')
    total_iva_emitido = Decimal('0')
    total_facturacion_emitida = Decimal('0')
    tipo_iva_decimal = Decimal('21')
    facturas_emitidas = []
    for p in presupuestos_emitidos:
        base, iva, total, tipo_iva = _calcular_totales_presupuesto(p)
        total_base_emitida += base
        total_iva_emitido += iva
        total_facturacion_emitida += total
        fecha_ref = p.fd_deliverynote_sent_at or p.fecha_creacion
        facturas_emitidas.append({
            'serie': 'FD',
            'numero': p.numero_solicitud or str(p.id),
            'fecha_expedicion': fecha_ref,
            'nombre': p.cliente.nombre if p.cliente else 'N/A',
            'importe_total': total,
            'base_calc': base,
            'iva_calc': iva,
            'tipo_iva': tipo_iva
        })
    
    total_base_soportada = sum(Decimal(str(f.base_imponible)) for f in facturas_proveedor)
    total_iva_soportado = sum(Decimal(str(f.importe_iva)) for f in facturas_proveedor)
    total_facturacion_soportada = sum(Decimal(str(f.total)) for f in facturas_proveedor)
    
    return render_template('informes/detalle_iva.html',
                         facturas_emitidas=facturas_emitidas,
                         facturas_proveedor=facturas_proveedor,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         tipo_iva_decimal=tipo_iva_decimal,
                         total_base_emitida=total_base_emitida,
                         total_iva_emitido=total_iva_emitido,
                         total_facturacion_emitida=total_facturacion_emitida,
                         total_base_soportada=total_base_soportada,
                         total_iva_soportado=total_iva_soportado,
                         total_facturacion_soportada=total_facturacion_soportada)

