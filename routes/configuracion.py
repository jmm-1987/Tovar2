"""Rutas de configuración (solo supervisor)"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, current_app
from flask_login import login_required, current_user
from extensions import db
from models import Usuario, Comercial, Cliente, Prenda, Pedido, LineaPedido, Presupuesto, LineaPresupuesto, Ticket, LineaTicket, Factura, LineaFactura, PlantillaEmail
from utils.auth import supervisor_required
from datetime import datetime
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
import os
import shutil

configuracion_bp = Blueprint('configuracion', __name__)

@configuracion_bp.route('/configuracion')
@login_required
@supervisor_required
def index():
    """Panel de configuración principal"""
    usuarios = Usuario.query.all()
    return render_template('configuracion/index.html', usuarios=usuarios)

@configuracion_bp.route('/configuracion/usuarios', methods=['GET', 'POST'])
@login_required
@supervisor_required
def gestion_usuarios():
    """Gestión de usuarios"""
    if request.method == 'POST':
        try:
            usuario = request.form.get('usuario')
            password = request.form.get('password')
            correo = request.form.get('correo')
            telefono = request.form.get('telefono', '')
            rol = request.form.get('rol')
            
            if not usuario or not password or not correo or not rol:
                flash('Todos los campos son obligatorios', 'error')
                return redirect(url_for('configuracion.gestion_usuarios'))
            
            # Verificar si el usuario ya existe
            if Usuario.query.filter_by(usuario=usuario).first():
                flash('El usuario ya existe', 'error')
                return redirect(url_for('configuracion.gestion_usuarios'))
            
            nuevo_usuario = Usuario(
                usuario=usuario,
                correo=correo,
                telefono=telefono,
                rol=rol,
                activo=True
            )
            nuevo_usuario.set_password(password)
            
            db.session.add(nuevo_usuario)
            db.session.flush()  # Para obtener el ID del usuario
            
            # Si el rol es comercial o administracion, crear registro en comerciales
            if rol in ['comercial', 'administracion']:
                comercial = Comercial(usuario_id=nuevo_usuario.id, _nombre=nuevo_usuario.usuario)
                db.session.add(comercial)
            
            db.session.commit()
            
            flash('Usuario creado correctamente', 'success')
            return redirect(url_for('configuracion.gestion_usuarios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear usuario: {str(e)}', 'error')
    
    usuarios = Usuario.query.all()
    return render_template('configuracion/usuarios.html', usuarios=usuarios)

@configuracion_bp.route('/configuracion/usuarios/<int:id>/editar', methods=['POST'])
@login_required
@supervisor_required
def editar_usuario(id):
    """Editar usuario"""
    usuario = Usuario.query.get_or_404(id)
    rol_anterior = usuario.rol
    
    try:
        usuario.correo = request.form.get('correo')
        usuario.telefono = request.form.get('telefono', '')
        nuevo_rol = request.form.get('rol')
        usuario.rol = nuevo_rol
        
        # Si se proporciona nueva contraseña, actualizarla
        nueva_password = request.form.get('password')
        if nueva_password:
            usuario.set_password(nueva_password)
        
        # Gestionar comercial según el rol
        comercial_existente = Comercial.query.filter_by(usuario_id=usuario.id).first()
        
        if nuevo_rol in ['comercial', 'administracion']:
            # Si cambió a comercial o administracion y no tiene comercial, crearlo
            if not comercial_existente:
                comercial = Comercial(usuario_id=usuario.id, _nombre=usuario.usuario)
                db.session.add(comercial)
        else:
            # Si cambió a otro rol y tiene comercial, eliminarlo
            if comercial_existente:
                db.session.delete(comercial_existente)
        
        db.session.commit()
        flash('Usuario actualizado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar usuario: {str(e)}', 'error')
    
    return redirect(url_for('configuracion.gestion_usuarios'))

@configuracion_bp.route('/configuracion/usuarios/<int:id>/eliminar', methods=['POST'])
@login_required
@supervisor_required
def eliminar_usuario(id):
    """Eliminar usuario (desactivar)"""
    usuario = Usuario.query.get_or_404(id)
    
    # No permitir eliminar al supervisor actual
    if usuario.id == current_user.id:
        flash('No puedes desactivar tu propio usuario', 'error')
        return redirect(url_for('configuracion.gestion_usuarios'))
    
    try:
        usuario.activo = False
        db.session.commit()
        flash('Usuario desactivado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al desactivar usuario: {str(e)}', 'error')
    
    return redirect(url_for('configuracion.gestion_usuarios'))

@configuracion_bp.route('/configuracion/verifactu')
@login_required
@supervisor_required
def verifactu_info():
    """Información sobre API Verifactu"""
    return render_template('configuracion/verifactu.html')

@configuracion_bp.route('/configuracion/exportar')
@login_required
@supervisor_required
def exportar_bd():
    """Exportar base de datos"""
    formato = request.args.get('formato', 'excel')
    
    if formato == 'excel':
        return exportar_excel()
    elif formato == 'txt':
        return exportar_txt()
    else:
        flash('Formato no válido', 'error')
        return redirect(url_for('configuracion.index'))

def exportar_excel():
    """Exportar base de datos a Excel"""
    wb = Workbook()
    
    # Lista de modelos y sus nombres de hoja
    modelos = [
        ('Comerciales', Comercial),
        ('Clientes', Cliente),
        ('Prendas', Prenda),
        ('Pedidos', Pedido),
        ('LineasPedido', LineaPedido),
        ('Presupuestos', Presupuesto),
        ('LineasPresupuesto', LineaPresupuesto),
        ('Tickets', Ticket),
        ('LineasTicket', LineaTicket),
        ('Facturas', Factura),
        ('LineasFactura', LineaFactura),
        ('Usuarios', Usuario),
    ]
    
    for nombre_hoja, modelo in modelos:
        ws = wb.create_sheet(title=nombre_hoja)
        
        # Obtener todos los registros
        registros = modelo.query.all()
        
        if not registros:
            continue
        
        # Obtener columnas del modelo
        columnas = [col.name for col in modelo.__table__.columns]
        
        # Escribir encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for idx, col in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=idx, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos
        for row_idx, registro in enumerate(registros, 2):
            for col_idx, col in enumerate(columnas, 1):
                valor = getattr(registro, col)
                if isinstance(valor, datetime):
                    valor = valor.strftime('%Y-%m-%d %H:%M:%S')
                elif valor is None:
                    valor = ''
                ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_bd_{fecha}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def exportar_txt():
    """Exportar base de datos a TXT (CSV)"""
    output = io.StringIO()
    
    modelos = [
        ('Comerciales', Comercial),
        ('Clientes', Cliente),
        ('Prendas', Prenda),
        ('Pedidos', Pedido),
        ('LineasPedido', LineaPedido),
        ('Presupuestos', Presupuesto),
        ('LineasPresupuesto', LineaPresupuesto),
        ('Tickets', Ticket),
        ('LineasTicket', LineaTicket),
        ('Facturas', Factura),
        ('LineasFactura', LineaFactura),
        ('Usuarios', Usuario),
    ]
    
    for nombre_tabla, modelo in modelos:
        output.write(f'\n{"="*80}\n')
        output.write(f'TABLA: {nombre_tabla}\n')
        output.write(f'{"="*80}\n\n')
        
        registros = modelo.query.all()
        
        if not registros:
            output.write('(Sin registros)\n\n')
            continue
        
        # Obtener columnas
        columnas = [col.name for col in modelo.__table__.columns]
        
        # Escribir encabezados
        writer = csv.writer(output)
        writer.writerow(columnas)
        
        # Escribir datos
        for registro in registros:
            fila = []
            for col in columnas:
                valor = getattr(registro, col)
                if isinstance(valor, datetime):
                    valor = valor.strftime('%Y-%m-%d %H:%M:%S')
                elif valor is None:
                    valor = ''
                fila.append(str(valor))
            writer.writerow(fila)
        
        output.write('\n')
    
    output.seek(0)
    
    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'backup_bd_{fecha}.txt'
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name=filename
    )

@configuracion_bp.route('/configuracion/importar', methods=['GET', 'POST'])
@login_required
@supervisor_required
def importar_bd():
    """Importar base de datos desde archivo"""
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('configuracion.importar_bd'))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('configuracion.importar_bd'))
        
        formato = request.form.get('formato', 'excel')
        
        try:
            if formato == 'excel':
                flash('La importación desde Excel está en desarrollo', 'info')
            elif formato == 'txt':
                flash('La importación desde TXT está en desarrollo', 'info')
            else:
                flash('Formato no válido', 'error')
        except Exception as e:
            flash(f'Error al importar: {str(e)}', 'error')
        
        return redirect(url_for('configuracion.importar_bd'))
    
    return render_template('configuracion/importar.html')

@configuracion_bp.route('/configuracion/plantillas-email')
@login_required
@supervisor_required
def plantillas_email():
    """Gestión de plantillas de email"""
    # Filtrar para excluir la plantilla genérica 'cambio_estado_pedido'
    plantillas = PlantillaEmail.query.filter(
        PlantillaEmail.tipo != 'cambio_estado_pedido'
    ).all()
    return render_template('configuracion/plantillas_email.html', plantillas=plantillas)

@configuracion_bp.route('/configuracion/plantillas-email/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@supervisor_required
def editar_plantilla_email(id):
    """Editar plantilla de email"""
    plantilla = PlantillaEmail.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            plantilla.asunto = request.form.get('asunto', '')
            plantilla.cuerpo = request.form.get('cuerpo', '')
            # Checkbox: si viene marcado es 'on', si no viene es None
            plantilla.enviar_activo = request.form.get('enviar_activo') == 'on'
            plantilla.fecha_actualizacion = datetime.utcnow()
            
            db.session.commit()
            flash('Plantilla actualizada correctamente', 'success')
            return redirect(url_for('configuracion.plantillas_email'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar plantilla: {str(e)}', 'error')
    
    return render_template('configuracion/editar_plantilla_email.html', plantilla=plantilla)

@configuracion_bp.route('/configuracion/plantillas-email/<int:id>/toggle', methods=['POST'])
@login_required
@supervisor_required
def toggle_plantilla_email(id):
    """Activar/desactivar el envío de una plantilla de email"""
    plantilla = PlantillaEmail.query.get_or_404(id)
    try:
        # Invertir el estado actual
        plantilla.enviar_activo = not plantilla.enviar_activo
        plantilla.fecha_actualizacion = datetime.utcnow()
        db.session.commit()
        
        estado = 'activada' if plantilla.enviar_activo else 'desactivada'
        flash(f'Plantilla {estado} correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar el estado de la plantilla: {str(e)}', 'error')
    
    return redirect(url_for('configuracion.plantillas_email'))

@configuracion_bp.route('/configuracion/descargar-bd')
@login_required
@supervisor_required
def descargar_bd():
    """Descargar el archivo de base de datos SQLite"""
    try:
        # Obtener la ruta de la base de datos desde la configuración
        database_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
        
        # Extraer la ruta del archivo desde sqlite:///ruta
        if database_uri.startswith('sqlite:///'):
            db_path = database_uri.replace('sqlite:///', '')
            # En Windows, puede venir con barras normales, convertir a barras del sistema
            db_path = os.path.normpath(db_path)
        else:
            flash('No se pudo determinar la ruta de la base de datos', 'error')
            return redirect(url_for('configuracion.index'))
        
        # Verificar que el archivo existe
        if not os.path.exists(db_path):
            flash('El archivo de base de datos no existe', 'error')
            return redirect(url_for('configuracion.index'))
        
        # Crear una copia temporal con nombre con fecha
        fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'pedidos_backup_{fecha}.db'
        
        # Crear un archivo temporal en memoria
        temp_file = io.BytesIO()
        with open(db_path, 'rb') as f:
            temp_file.write(f.read())
        temp_file.seek(0)
        
        return send_file(
            temp_file,
            mimetype='application/x-sqlite3',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        flash(f'Error al descargar la base de datos: {str(e)}', 'error')
        return redirect(url_for('configuracion.index'))

@configuracion_bp.route('/configuracion/importar-bd-sqlite', methods=['GET', 'POST'])
@login_required
@supervisor_required
def importar_bd_sqlite():
    """Importar/cargar un archivo SQLite para reemplazar la base de datos actual"""
    if request.method == 'POST':
        try:
            # Verificar que se subió un archivo
            if 'archivo' not in request.files:
                flash('No se seleccionó ningún archivo', 'error')
                return redirect(url_for('configuracion.importar_bd_sqlite'))
            
            archivo = request.files['archivo']
            if archivo.filename == '':
                flash('No se seleccionó ningún archivo', 'error')
                return redirect(url_for('configuracion.importar_bd_sqlite'))
            
            # Verificar extensión
            if not archivo.filename.lower().endswith('.db'):
                flash('El archivo debe ser un archivo SQLite (.db)', 'error')
                return redirect(url_for('configuracion.importar_bd_sqlite'))
            
            # Obtener la ruta de la base de datos actual
            database_uri = current_app.config['SQLALCHEMY_DATABASE_URI']
            
            if not database_uri.startswith('sqlite:///'):
                flash('No se pudo determinar la ruta de la base de datos', 'error')
                return redirect(url_for('configuracion.importar_bd_sqlite'))
            
            db_path = database_uri.replace('sqlite:///', '')
            db_path = os.path.normpath(db_path)
            
            # Validar que el archivo subido es SQLite válido (verificar header)
            archivo.seek(0)
            header = archivo.read(16)
            archivo.seek(0)
            
            # SQLite tiene un header específico: "SQLite format 3\000"
            if not header.startswith(b'SQLite format 3\x00'):
                flash('El archivo no es un archivo SQLite válido', 'error')
                return redirect(url_for('configuracion.importar_bd_sqlite'))
            
            # Cerrar todas las conexiones de la base de datos antes de reemplazar
            db.session.close()
            db.engine.dispose()
            
            # Hacer backup de la base de datos actual antes de reemplazarla
            if os.path.exists(db_path):
                fecha_backup = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_path = db_path.replace('.db', f'_backup_{fecha_backup}.db')
                try:
                    shutil.copy2(db_path, backup_path)
                    flash(f'Backup creado: {os.path.basename(backup_path)}', 'info')
                except Exception as e:
                    flash(f'Advertencia: No se pudo crear backup automático: {str(e)}', 'warning')
            
            # Guardar el nuevo archivo
            archivo.save(db_path)
            
            # Reinicializar la conexión de la base de datos
            db.engine.dispose()
            db.create_all()
            
            flash('Base de datos importada correctamente. La aplicación se reiniciará.', 'success')
            return redirect(url_for('configuracion.index'))
            
        except Exception as e:
            flash(f'Error al importar la base de datos: {str(e)}', 'error')
            import traceback
            traceback.print_exc()
            return redirect(url_for('configuracion.importar_bd_sqlite'))
    
    return render_template('configuracion/importar_bd_sqlite.html')

