"""Rutas para informes y reportes"""
from flask import Blueprint, render_template, request, jsonify, send_file, make_response, redirect, url_for
from flask_login import login_required
from datetime import datetime
from decimal import Decimal
from extensions import db
from models import Factura, FacturaProveedor, Nomina, Empleado, LineaFactura, Cliente, Ticket, OtroGasto
from sqlalchemy import func, extract, not_
from utils.auth import not_usuario_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from playwright.sync_api import sync_playwright
import io
import tempfile
import os

informes_bp = Blueprint('informes', __name__)

@informes_bp.route('/informes')
@login_required
@not_usuario_required
def index():
    """Página principal de informes"""
    return render_template('informes/index.html')

@informes_bp.route('/informes/facturacion-emitida')
@login_required
@not_usuario_required
def facturacion_emitida():
    """Informe de facturación emitida con filtros por mes o trimestre"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')  # 'mes' o 'trimestre'
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    # Base query - Excluir albaranes (formato A2601_XXX)
    query = Factura.query.filter(
        extract('year', Factura.fecha_expedicion) == año,
        not_(Factura.numero.like('A%_%'))
    )
    
    # Aplicar filtro por cliente si está seleccionado
    if cliente_id:
        query = query.filter(Factura.cliente_id == cliente_id)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', Factura.fecha_expedicion) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', Factura.fecha_expedicion).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(Factura.fecha_expedicion.desc(), Factura.numero.desc()).all()
    
    # Calcular totales
    total_facturacion = sum(f.importe_total for f in facturas)
    total_iva_repercutido = Decimal('0')
    tipo_iva = Decimal('21')  # IVA estándar al 21%
    
    # Calcular IVA repercutido desde las líneas de factura
    # Asumimos que el importe incluye IVA al 21%
    for factura in facturas:
        for linea in factura.lineas:
            importe_con_iva = Decimal(str(linea.importe))
            # Calcular base imponible: importe / (1 + tipo_iva/100)
            base_imponible = importe_con_iva / (Decimal('1') + tipo_iva / Decimal('100'))
            # Calcular IVA repercutido: base_imponible * (tipo_iva/100)
            iva_linea = base_imponible * tipo_iva / Decimal('100')
            total_iva_repercutido += iva_linea.quantize(Decimal('0.01'))
    
    # Obtener lista de clientes para el filtro
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    cliente_seleccionado = Cliente.query.get(cliente_id) if cliente_id else None
    
    return render_template('informes/facturacion_emitida.html', 
                         facturas=facturas,
                         total_facturacion=total_facturacion,
                         total_iva_repercutido=total_iva_repercutido,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         clientes=clientes,
                         cliente_id=cliente_id,
                         cliente_seleccionado=cliente_seleccionado)

@informes_bp.route('/informes/facturacion-emitida/exportar-excel')
@login_required
@not_usuario_required
def facturacion_emitida_exportar_excel():
    """Exportar informe de facturación emitida a Excel"""
    # Obtener parámetros de filtro (mismos que en el informe principal)
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    # Base query - Excluir albaranes (formato A2601_XXX)
    query = Factura.query.filter(
        extract('year', Factura.fecha_expedicion) == año,
        not_(Factura.numero.like('A%_%'))
    )
    
    # Aplicar filtro por cliente si está seleccionado
    if cliente_id:
        query = query.filter(Factura.cliente_id == cliente_id)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', Factura.fecha_expedicion) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', Factura.fecha_expedicion).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(Factura.fecha_expedicion.desc(), Factura.numero.desc()).all()
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturación Emitida"
    
    # Estilos para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir encabezados
    headers = ['DNI/CIF', 'Nombre Cliente', 'Número Factura', 'Fecha', 'Base Imponible', 'IVA', 'Total', 'Estado Cobro']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Escribir datos
    for row_idx, factura in enumerate(facturas, 2):
        # Obtener DNI/CIF (de factura.nif o factura.cliente.nif)
        dni_cif = factura.nif if factura.nif else ''
        if not dni_cif and factura.cliente:
            dni_cif = factura.cliente.nif if factura.cliente.nif else ''
        
        # Nombre del cliente
        nombre_cliente = factura.nombre if factura.nombre else ''
        if not nombre_cliente and factura.cliente:
            nombre_cliente = factura.cliente.nombre if factura.cliente.nombre else ''
        
        # Número de factura
        numero_factura = f"{factura.serie}-{factura.numero}"
        
        # Fecha de expedición
        fecha_expedicion = factura.fecha_expedicion.strftime('%d/%m/%Y') if factura.fecha_expedicion else ''
        
        # Calcular base imponible e IVA
        tipo_iva_val = float(factura.tipo_iva) if factura.tipo_iva else 21.0
        importe_total_val = float(factura.importe_total) if factura.importe_total else 0.0
        descuento_pronto_pago_val = float(factura.descuento_pronto_pago) if factura.descuento_pronto_pago else 0.0
        
        if descuento_pronto_pago_val > 0:
            # Si hay descuento por pronto pago, el importe_total ya lo tiene aplicado
            subtotal_sin_descuento = importe_total_val / (1 - descuento_pronto_pago_val / 100)
            base_imponible = subtotal_sin_descuento / (1 + tipo_iva_val / 100)
            importe_iva = subtotal_sin_descuento - base_imponible
        else:
            # Sin descuento: importe_total = base + iva = base * (1 + tipo_iva/100)
            base_imponible = importe_total_val / (1 + tipo_iva_val / 100)
            importe_iva = importe_total_val - base_imponible
        
        # Estado del cobro
        estado_cobro = factura.estado_cobro if factura.estado_cobro else 'pendiente'
        estado_cobro_texto = {
            'pendiente': 'Pendiente',
            'cobrada_parcialmente': 'Cobrada Parcialmente',
            'cobrada': 'Cobrada'
        }.get(estado_cobro, estado_cobro.title())
        
        # Escribir fila
        ws.cell(row=row_idx, column=1, value=dni_cif)
        ws.cell(row=row_idx, column=2, value=nombre_cliente)
        ws.cell(row=row_idx, column=3, value=numero_factura)
        ws.cell(row=row_idx, column=4, value=fecha_expedicion)
        ws.cell(row=row_idx, column=5, value=round(base_imponible, 2))
        ws.cell(row=row_idx, column=6, value=round(importe_iva, 2))
        ws.cell(row=row_idx, column=7, value=round(importe_total_val, 2))
        ws.cell(row=row_idx, column=8, value=estado_cobro_texto)
        
        # Formato numérico para columnas de importes
        ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'
    
    # Ajustar ancho de columnas automáticamente según el contenido
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # Para números, calcular longitud considerando formato
                    if isinstance(cell.value, (int, float)):
                        cell_length = len(f"{cell.value:,.2f}")
                    else:
                        cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        # Añadir un poco de espacio extra y establecer límites mínimos y máximos
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Ajustar altura de fila de encabezado
    ws.row_dimensions[1].height = 25
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo con el período
    filename = f'facturacion_emitida_{periodo_label.replace(" ", "_")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@informes_bp.route('/informes/facturacion-emitida/exportar-pdf')
@login_required
@not_usuario_required
def facturacion_emitida_exportar_pdf():
    """Exportar informe de facturación emitida a PDF (vertical, una hoja)"""
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    query = Factura.query.filter(
        extract('year', Factura.fecha_expedicion) == año,
        not_(Factura.numero.like('A%_%'))
    )
    if cliente_id:
        query = query.filter(Factura.cliente_id == cliente_id)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', Factura.fecha_expedicion) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', Factura.fecha_expedicion).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(Factura.fecha_expedicion.desc(), Factura.numero.desc()).all()
    
    # Calcular totales
    total_facturacion = sum(f.importe_total for f in facturas)
    total_base = Decimal('0')
    total_iva = Decimal('0')
    for factura in facturas:
        tipo_iva_val = float(factura.tipo_iva) if factura.tipo_iva else 21.0
        importe_total_val = float(factura.importe_total) if factura.importe_total else 0.0
        descuento_val = float(factura.descuento_pronto_pago) if factura.descuento_pronto_pago else 0.0
        if descuento_val > 0:
            subtotal_sin_descuento = importe_total_val / (1 - descuento_val / 100)
            base = subtotal_sin_descuento / (1 + tipo_iva_val / 100)
            iva = subtotal_sin_descuento - base
        else:
            base = importe_total_val / (1 + tipo_iva_val / 100)
            iva = importe_total_val - base
        total_base += Decimal(str(round(base, 2)))
        total_iva += Decimal(str(round(iva, 2)))
    
    total_iva_repercutido = total_iva
    
    html = render_template('informes/facturacion_emitida_pdf.html',
        facturas=facturas,
        periodo_label=periodo_label,
        total_facturacion=total_facturacion,
        total_iva_repercutido=total_iva_repercutido,
        total_base=total_base,
        total_iva=total_iva
    )
    
    pdf_buffer = BytesIO()
    try:
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as temp_file:
            temp_file.write(html)
            temp_html_path = temp_file.name
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(f'file://{temp_html_path}')
            pdf_bytes = page.pdf(
                format='A4',
                print_background=True,
                margin={'top': '6mm', 'right': '6mm', 'bottom': '6mm', 'left': '6mm'}
            )
            browser.close()
        
        pdf_buffer.write(pdf_bytes)
        try:
            os.unlink(temp_html_path)
        except Exception:
            pass
    except Exception as e:
        import traceback
        from flask import flash
        print(traceback.format_exc())
        flash(f'Error al generar PDF: {str(e)}', 'error')
        return redirect(url_for('informes.facturacion_emitida'))
    
    pdf_buffer.seek(0)
    filename = f'facturacion_emitida_{periodo_label.replace(" ", "_")}.pdf'
    
    response = make_response(pdf_buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename={filename}'
    return response

@informes_bp.route('/informes/facturacion-soportada')
@login_required
@not_usuario_required
def facturacion_soportada():
    """Informe de facturación soportada (facturas de proveedor) con filtros por mes o trimestre"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')  # 'mes' o 'trimestre'
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    # Base query
    query = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(FacturaProveedor.fecha_factura.desc()).all()
    
    # Calcular totales
    total_facturacion = sum(f.total for f in facturas)
    total_base_imponible = sum(f.base_imponible for f in facturas)
    total_iva_soportado = sum(f.importe_iva for f in facturas)
    
    return render_template('informes/facturacion_soportada.html',
                         facturas=facturas,
                         total_facturacion=total_facturacion,
                         total_base_imponible=total_base_imponible,
                         total_iva_soportado=total_iva_soportado,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label)

@informes_bp.route('/informes/nominas')
@login_required
@not_usuario_required
def nominas():
    """Informe de nóminas por empleado y global"""
    # Obtener parámetros de filtro
    empleado_id = request.args.get('empleado_id', None, type=int)
    año = request.args.get('año', datetime.now().year, type=int)
    
    # Base query
    query = Nomina.query.filter(Nomina.año == año)
    
    if empleado_id:
        query = query.filter(Nomina.empleado_id == empleado_id)
    
    nominas = query.order_by(Nomina.mes.desc(), Nomina.empleado_id).all()
    
    # Calcular totales
    total_global = sum(n.total_devengado for n in nominas)
    
    # Totales por empleado
    totales_por_empleado = {}
    for nomina in nominas:
        empleado_nombre = nomina.empleado.nombre if nomina.empleado else 'Sin empleado'
        if empleado_nombre not in totales_por_empleado:
            totales_por_empleado[empleado_nombre] = Decimal('0')
        totales_por_empleado[empleado_nombre] += nomina.total_devengado
    
    # Obtener lista de empleados para el filtro
    empleados = Empleado.query.order_by(Empleado.nombre).all()
    
    return render_template('informes/nominas.html',
                         nominas=nominas,
                         total_global=total_global,
                         totales_por_empleado=totales_por_empleado,
                         empleados=empleados,
                         empleado_id=empleado_id,
                         año=año)

@informes_bp.route('/informes/iva')
@login_required
@not_usuario_required
def iva():
    """Informe de IVA: contrastar IVA repercutido vs IVA soportado"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')  # 'mes' o 'trimestre'
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    # Calcular IVA repercutido (de facturas emitidas) - Excluir albaranes
    query_facturas = Factura.query.filter(
        extract('year', Factura.fecha_expedicion) == año,
        not_(Factura.numero.like('A%_%'))
    )
    
    if tipo_filtro == 'mes' and periodo:
        query_facturas = query_facturas.filter(extract('month', Factura.fecha_expedicion) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query_facturas = query_facturas.filter(extract('month', Factura.fecha_expedicion).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas_emitidas = query_facturas.all()
    iva_repercutido = Decimal('0')
    base_repercutida = Decimal('0')
    tipo_iva = Decimal('21')  # IVA estándar al 21%
    
    for factura in facturas_emitidas:
        for linea in factura.lineas:
            importe_con_iva = Decimal(str(linea.importe))
            # Calcular base imponible: importe / (1 + tipo_iva/100)
            base_imponible = importe_con_iva / (Decimal('1') + tipo_iva / Decimal('100'))
            base_imponible = base_imponible.quantize(Decimal('0.01'))
            base_repercutida += base_imponible
            # Calcular IVA repercutido: base_imponible * (tipo_iva/100)
            iva_linea = base_imponible * tipo_iva / Decimal('100')
            iva_repercutido += iva_linea.quantize(Decimal('0.01'))
    
    # Calcular IVA soportado (de facturas de proveedor)
    query_facturas_prov = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
    
    facturas_proveedor = query_facturas_prov.all()
    iva_soportado = sum(f.importe_iva for f in facturas_proveedor)
    base_soportada = sum(f.base_imponible for f in facturas_proveedor)
    
    # Calcular diferencia
    diferencia_iva = iva_repercutido - iva_soportado
    
    return render_template('informes/iva.html',
                         iva_repercutido=iva_repercutido,
                         base_repercutida=base_repercutida,
                         iva_soportado=iva_soportado,
                         base_soportada=base_soportada,
                         diferencia_iva=diferencia_iva,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         num_facturas_emitidas=len(facturas_emitidas),
                         num_facturas_proveedor=len(facturas_proveedor))

@informes_bp.route('/informes/resultados')
@login_required
@not_usuario_required
def resultados():
    """Informe de resultados: ingresos - gastos categorizados por tipo, con gráficas mes a mes"""
    año = request.args.get('año', datetime.now().year, type=int)
    
    MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 
             'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    datos_meses = []
    
    for mes in range(1, 13):
        # Ingresos: Facturas emitidas (excl. albaranes) + Tickets tienda
        facturas = Factura.query.filter(
            extract('year', Factura.fecha_expedicion) == año,
            extract('month', Factura.fecha_expedicion) == mes,
            not_(Factura.numero.like('A%_%'))
        ).all()
        ingresos_facturas = sum(Decimal(str(f.importe_total)) for f in facturas)
        
        tickets = Ticket.query.filter(
            extract('year', Ticket.fecha_expedicion) == año,
            extract('month', Ticket.fecha_expedicion) == mes
        ).all()
        ingresos_tickets = sum(Decimal(str(t.importe_total)) for t in tickets)
        ingresos = ingresos_facturas + ingresos_tickets
        
        # Gastos por categoría
        facturas_prov = FacturaProveedor.query.filter(
            extract('year', FacturaProveedor.fecha_factura) == año,
            extract('month', FacturaProveedor.fecha_factura) == mes
        ).all()
        gasto_facturas = sum(Decimal(str(f.total)) for f in facturas_prov)
        
        nominas = Nomina.query.filter(Nomina.año == año, Nomina.mes == mes).all()
        gasto_nominas = sum(Decimal(str(n.total_devengado)) for n in nominas)
        
        recibo_iva = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'recibo_iva'
        ).all()
        gasto_iva = sum(Decimal(str(g.importe)) for g in recibo_iva)
        
        recibo_irpf = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'recibo_irpf'
        ).all()
        gasto_irpf = sum(Decimal(str(g.importe)) for g in recibo_irpf)
        
        seg_social = OtroGasto.query.filter(
            extract('year', OtroGasto.fecha) == año,
            extract('month', OtroGasto.fecha) == mes,
            OtroGasto.tipo == 'seguridad_social'
        ).all()
        gasto_seg_social = sum(Decimal(str(g.importe)) for g in seg_social)
        
        total_gastos = gasto_facturas + gasto_nominas + gasto_iva + gasto_irpf + gasto_seg_social
        resultado = ingresos - total_gastos
        
        datos_meses.append({
            'mes': mes,
            'mes_nombre': MESES[mes],
            'ingresos': float(ingresos),
            'gasto_facturas': float(gasto_facturas),
            'gasto_nominas': float(gasto_nominas),
            'gasto_iva': float(gasto_iva),
            'gasto_irpf': float(gasto_irpf),
            'gasto_seg_social': float(gasto_seg_social),
            'total_gastos': float(total_gastos),
            'resultado': float(resultado)
        })
    
    # Totales anuales
    total_ingresos = sum(d['ingresos'] for d in datos_meses)
    total_gastos_anual = sum(d['total_gastos'] for d in datos_meses)
    resultado_anual = total_ingresos - total_gastos_anual
    
    return render_template('informes/resultados.html',
                         datos_meses=datos_meses,
                         año=año,
                         total_ingresos=total_ingresos,
                         total_gastos_anual=total_gastos_anual,
                         resultado_anual=resultado_anual)

@informes_bp.route('/informes/facturacion-emitida/detalle')
@login_required
@not_usuario_required
def facturacion_emitida_detalle():
    """Detalle de facturación emitida con listado completo de facturas"""
    # Obtener parámetros de filtro (mismos que en el informe principal)
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    cliente_id = request.args.get('cliente_id', None, type=int)
    
    # Base query - Excluir albaranes (formato A2601_XXX)
    query = Factura.query.filter(
        extract('year', Factura.fecha_expedicion) == año,
        not_(Factura.numero.like('A%_%'))
    )
    
    # Aplicar filtro por cliente si está seleccionado
    if cliente_id:
        query = query.filter(Factura.cliente_id == cliente_id)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', Factura.fecha_expedicion) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', Factura.fecha_expedicion).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(Factura.fecha_expedicion.desc(), Factura.numero.desc()).all()
    
    # Obtener cliente seleccionado si existe
    cliente_seleccionado = Cliente.query.get(cliente_id) if cliente_id else None
    
    return render_template('informes/detalle_facturacion_emitida.html',
                         facturas=facturas,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         cliente_id=cliente_id,
                         cliente_seleccionado=cliente_seleccionado)

@informes_bp.route('/informes/facturacion-soportada/detalle')
@login_required
@not_usuario_required
def facturacion_soportada_detalle():
    """Detalle de facturación soportada con listado completo de facturas de proveedor"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    # Base query
    query = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query = query.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query = query.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas = query.order_by(FacturaProveedor.fecha_factura.desc()).all()
    
    return render_template('informes/detalle_facturacion_soportada.html',
                         facturas=facturas,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label)

@informes_bp.route('/informes/nominas/detalle')
@login_required
@not_usuario_required
def nominas_detalle():
    """Detalle de nóminas con listado completo"""
    # Obtener parámetros de filtro
    empleado_id = request.args.get('empleado_id', None, type=int)
    año = request.args.get('año', datetime.now().year, type=int)
    
    # Base query
    query = Nomina.query.filter(Nomina.año == año)
    
    if empleado_id:
        query = query.filter(Nomina.empleado_id == empleado_id)
    
    nominas = query.order_by(Nomina.mes.desc(), Nomina.empleado_id).all()
    
    # Obtener lista de empleados para el filtro
    empleados = Empleado.query.order_by(Empleado.nombre).all()
    empleado_seleccionado = Empleado.query.get(empleado_id) if empleado_id else None
    
    return render_template('informes/detalle_nominas.html',
                         nominas=nominas,
                         empleados=empleados,
                         empleado_id=empleado_id,
                         empleado_seleccionado=empleado_seleccionado,
                         año=año)

@informes_bp.route('/informes/iva/detalle')
@login_required
@not_usuario_required
def iva_detalle():
    """Detalle de IVA con listado completo de facturas emitidas y de proveedor"""
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'mes')
    año = request.args.get('año', datetime.now().year, type=int)
    periodo = request.args.get('periodo', None, type=int)
    
    # Calcular IVA repercutido (de facturas emitidas) - Excluir albaranes
    query_facturas = Factura.query.filter(
        extract('year', Factura.fecha_expedicion) == año,
        not_(Factura.numero.like('A%_%'))
    )
    
    if tipo_filtro == 'mes' and periodo:
        query_facturas = query_facturas.filter(extract('month', Factura.fecha_expedicion) == periodo)
        periodo_label = f"{['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][periodo]} {año}"
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query_facturas = query_facturas.filter(extract('month', Factura.fecha_expedicion).between(mes_inicio, mes_fin))
        trimestres = {1: '1T', 2: '2T', 3: '3T', 4: '4T'}
        periodo_label = f"{trimestres.get(periodo, '')} {año}"
    else:
        periodo_label = f"Año {año}"
    
    facturas_emitidas = query_facturas.order_by(Factura.fecha_expedicion.desc(), Factura.numero.desc()).all()
    
    # Calcular IVA soportado (de facturas de proveedor)
    query_facturas_prov = FacturaProveedor.query.filter(extract('year', FacturaProveedor.fecha_factura) == año)
    
    if tipo_filtro == 'mes' and periodo:
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura) == periodo)
    elif tipo_filtro == 'trimestre' and periodo:
        mes_inicio = (periodo - 1) * 3 + 1
        mes_fin = periodo * 3
        query_facturas_prov = query_facturas_prov.filter(extract('month', FacturaProveedor.fecha_factura).between(mes_inicio, mes_fin))
    
    facturas_proveedor = query_facturas_prov.order_by(FacturaProveedor.fecha_factura.desc()).all()
    
    # Precalcular totales para evitar problemas de tipo en las plantillas
    tipo_iva_decimal = Decimal('21')
    total_base_emitida = Decimal('0')
    total_iva_emitido = Decimal('0')
    total_facturacion_emitida = Decimal('0')
    
    for factura in facturas_emitidas:
        total_facturacion_emitida += Decimal(str(factura.importe_total))
        for linea in factura.lineas:
            importe_con_iva = Decimal(str(linea.importe))
            base_imponible = importe_con_iva / (Decimal('1') + tipo_iva_decimal / Decimal('100'))
            base_imponible = base_imponible.quantize(Decimal('0.01'))
            total_base_emitida += base_imponible
            iva_linea = base_imponible * tipo_iva_decimal / Decimal('100')
            total_iva_emitido += iva_linea.quantize(Decimal('0.01'))
    
    total_base_soportada = sum(Decimal(str(f.base_imponible)) for f in facturas_proveedor)
    total_iva_soportado = sum(Decimal(str(f.importe_iva)) for f in facturas_proveedor)
    total_facturacion_soportada = sum(Decimal(str(f.total)) for f in facturas_proveedor)
    
    return render_template('informes/detalle_iva.html',
                         facturas_emitidas=facturas_emitidas,
                         facturas_proveedor=facturas_proveedor,
                         tipo_filtro=tipo_filtro,
                         año=año,
                         periodo=periodo,
                         periodo_label=periodo_label,
                         tipo_iva_decimal=tipo_iva_decimal,
                         total_base_emitida=total_base_emitida,
                         total_iva_emitido=total_iva_emitido,
                         total_facturacion_emitida=total_facturacion_emitida,
                         total_base_soportada=total_base_soportada,
                         total_iva_soportado=total_iva_soportado,
                         total_facturacion_soportada=total_facturacion_soportada)

